from datetime import datetime

from openpyxl import load_workbook


def log_expense(file_path, expense_name, amount, date_str=None):
    print(f"Loading {file_path}...")
    try:
        wb = load_workbook(file_path)
    except PermissionError:
        return "❌ Error: The file is open. Please close Excel and try again."

    # 1. Parse the date first
    if date_str:
        try:
            date_val = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            date_val = datetime.now()
    else:
        date_val = datetime.now()

    # 2. Find the Correct Sheet based on the Date
    target_month_name = date_val.strftime("%B")  # e.g., "February"

    if target_month_name in wb.sheetnames:
        sheet = wb[target_month_name]
        print(
            f"📅 Date is in {target_month_name}. Logging to sheet: '{target_month_name}'"
        )
    else:
        # Fallback: If "February" doesn't exist, use the most recent tab (Index 0)
        sheet = wb.worksheets[0]
        print(
            f"⚠️ Sheet '{target_month_name}' not found. Defaulting to: '{sheet.title}'"
        )

    # 3. Find the first empty row in the Expenses table (Columns B, C, D)
    target_row = None
    for row_idx in range(4, 100):
        cell_b = sheet.cell(row=row_idx, column=2)  # Column B

        if cell_b.value is None or str(cell_b.value).strip() == "":
            target_row = row_idx
            break

    if target_row:
        # Write Data
        sheet.cell(row=target_row, column=2, value=expense_name)
        sheet.cell(row=target_row, column=3, value=float(amount))
        sheet.cell(row=target_row, column=4, value=date_val)

        # Copy formatting from row above
        if target_row > 4:
            source_row = target_row - 1
            for col in [2, 3, 4]:
                source_cell = sheet.cell(row=source_row, column=col)
                target_cell = sheet.cell(row=target_row, column=col)
                if source_cell.has_style:
                    target_cell._style = source_cell._style

        wb.save(file_path)
        return f"✅ Logged: {expense_name} | ${amount} | {date_val.strftime('%Y-%m-%d')} in tab '{sheet.title}'"
    else:
        return "❌ Error: Could not find an empty row."


if __name__ == "__main__":
    # Test path
    path = "/Users/ramirolb/Library/CloudStorage/OneDrive-Personal/Excel Documents/monthly-expenses.xlsx"
    # Test adding a February expense
    print(log_expense(path, "Test Feb Item", 99.99, "2026-02-17"))
