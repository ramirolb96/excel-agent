import calendar
import re
import sys
from copy import copy
from datetime import datetime, timedelta

from openpyxl import load_workbook

# --- CONFIGURATION ---
# Fixed expenses that ALWAYS stay the same
FIXED_EXPENSES = ["Car Insurance", "Rent", "Car Payment"]
WEEKLY_INCOME_AMOUNT = 1040.64
# ---------------------


def get_fridays(year, month):
    """Returns a list of datetime objects for every Friday in the given month."""
    fridays = []
    cal = calendar.Calendar(firstweekday=calendar.SUNDAY)
    for day in cal.itermonthdates(year, month):
        if day.month == month and day.weekday() == 4:  # 4 = Friday
            fridays.append(day)
    return fridays


def copy_style(source_cell, target_cell):
    """Copies visual style from one cell to another."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)


def increment_installment(text):
    """
    Parses strings like 'Quince (1/6)' -> returns ('Quince (2/6)', False)
    If it was 'Quince (6/6)' -> returns (None, True) (indicating it's finished)
    """
    # Regex to find (X/Y) pattern
    match = re.search(r"\((\d+)/(\d+)\)", text)
    if match:
        current_val = int(match.group(1))
        total_val = int(match.group(2))

        if current_val < total_val:
            # Increment: (1/6) -> (2/6)
            new_val = current_val + 1
            new_text = text.replace(
                f"({current_val}/{total_val})", f"({new_val}/{total_val})"
            )
            return new_text, False  # False = Not Finished
        else:
            # Finished: (6/6) -> Remove
            return None, True  # True = Finished
    return text, False


def add_next_month(file_path):
    print(f"Loading {file_path}...")
    try:
        wb = load_workbook(file_path)
    except PermissionError:
        print("❌ Error: The file is open. Please close Excel and try again.")
        return

    # 1. Determine Target Month
    now = datetime.now()
    current_month_name = now.strftime("%B")
    latest_sheet = wb.worksheets[0]

    # Decide if we are creating NEXT month or CURRENT month
    if current_month_name == latest_sheet.title:
        target_date = now.replace(day=1) + timedelta(days=32)
        target_date = target_date.replace(day=1)
    else:
        target_date = now

    target_month_name = target_date.strftime("%B")
    target_year = target_date.year
    target_month_num = target_date.month

    if target_month_name in wb.sheetnames:
        print(f"⚠️  Stop: Sheet '{target_month_name}' already exists.")
        return

    print(f"Creating '{target_month_name}' (Year: {target_year})...")

    # 2. Copy Template Sheet
    new_sheet = wb.copy_worksheet(latest_sheet)
    new_sheet.title = target_month_name
    wb.move_sheet(new_sheet, offset=-(len(wb.sheetnames) - 1))

    # --- PHASE 1: INCOME AUTOMATION (Right Side) ---
    fridays = get_fridays(target_year, target_month_num)
    print(f"💰 Found {len(fridays)} paydays. Setting income...")

    # Identify Week Rows
    week_rows = {}
    last_week_row_idx = 0
    for row in new_sheet.iter_rows(min_col=12, max_col=12):
        cell = row[0]
        if isinstance(cell.value, str) and cell.value.startswith("Week "):
            week_rows[cell.value] = cell.row
            last_week_row_idx = max(last_week_row_idx, cell.row)

    # Process Weeks 1-4
    for i, friday in enumerate(fridays):
        week_label = f"Week {i + 1}"

        if week_label in week_rows:
            r = week_rows[week_label]
        elif i == 4:  # Add Week 5 if needed
            r = last_week_row_idx + 1
            print("➕  Adding row for Week 5...")
            new_sheet.cell(row=r, column=12, value="Week 5")
            ref_r = week_rows.get("Week 4", r - 1)
            for c in [12, 13, 14]:
                copy_style(
                    new_sheet.cell(row=ref_r, column=c), new_sheet.cell(row=r, column=c)
                )
        else:
            continue

        # Set Amount & Date
        new_sheet.cell(row=r, column=13, value=WEEKLY_INCOME_AMOUNT)
        new_sheet.cell(row=r, column=14, value=friday)

    # Remove extra Week 5 if not needed
    if len(fridays) < 5 and "Week 5" in week_rows:
        print("➖  Removing extra Week 5...")
        r = week_rows["Week 5"]
        for c in [12, 13, 14]:
            new_sheet.cell(row=r, column=c, value=None)

    # --- PHASE 2: EXPENSE AUTOMATION (Left Side) ---
    print("🧾 Processing expenses...")

    # Iterate through Expense Table (Cols B, C, D)
    for row in new_sheet.iter_rows(min_col=2, max_col=4):
        name_cell = row[0]  # Col B
        amt_cell = row[1]  # Col C
        date_cell = row[2]  # Col D

        if not name_cell.value or str(name_cell.value) == "Expense":
            continue

        expense_name = str(name_cell.value).strip()

        # LOGIC 1: Handle Installments (e.g. "Quince (1/6)")
        new_name, is_finished = increment_installment(expense_name)

        if is_finished:
            # Case A: Debt Paid Off -> Clear Entire Row
            print(f"🎉  Finished payment: {expense_name}. Removing from list.")
            name_cell.value = None
            amt_cell.value = None
            date_cell.value = None
            continue

        if new_name != expense_name:
            # Case B: Installment continues -> Update Name, Keep Amount & Date Logic
            print(f"🔄  Updating installment: {expense_name} -> {new_name}")
            name_cell.value = new_name
            # Installments act like Fixed Expenses (Keep Amount, Update Date)
            expense_name = new_name  # Treat as valid for next check

            # Update Date
            if isinstance(date_cell.value, datetime):
                try:
                    new_date = date_cell.value.replace(
                        year=target_year, month=target_month_num
                    )
                    date_cell.value = new_date
                except ValueError:
                    date_cell.value = None
            continue

        # LOGIC 2: Handle Fixed Expenses
        if expense_name in FIXED_EXPENSES:
            # Keep Amount, Update Date
            if isinstance(date_cell.value, datetime):
                try:
                    new_date = date_cell.value.replace(
                        year=target_year, month=target_month_num
                    )
                    date_cell.value = new_date
                except ValueError:
                    date_cell.value = None
        else:
            # LOGIC 3: Variable Expenses -> Clear Amount & Date
            if isinstance(amt_cell.value, (int, float)):
                amt_cell.value = None
            if isinstance(date_cell.value, (datetime, int, float)):
                date_cell.value = None

    # Save
    try:
        wb.save(file_path)
        print(f"✅ Success! Created '{target_month_name}'")
    except PermissionError:
        print("❌ Error: Could not save. Close the file and try again.")


if __name__ == "__main__":
    default_path = "/Users/ramirolb/Library/CloudStorage/OneDrive-Personal/Excel Documents/monthly-expenses.xlsx"
    target_file = sys.argv[1] if len(sys.argv) > 1 else default_path
    add_next_month(target_file)
